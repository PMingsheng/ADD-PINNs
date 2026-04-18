import argparse
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.spatial import cKDTree

from config import DataConfig, SAMPLING_CONFIGS


def _extract_epoch_from_name(name: str) -> int:
    patterns = [
        r"^(?:u|f)_heatmap_(\d+)\.npz$",
        r"^phi_epoch_(\d+)\.npz$",
        r"^epoch_(\d+)$",
    ]
    for pat in patterns:
        match = re.match(pat, name)
        if match:
            return int(match.group(1))
    return -1


def find_latest_heatmap_npz(root_dir: Path, stem: str) -> Optional[Path]:
    candidates = list(root_dir.glob(f"outputs_flower/*/{stem}_heatmaps/{stem}_heatmap_*.npz"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: (_extract_epoch_from_name(p.name), p.stat().st_mtime))


def infer_epoch_from_npz_path(npz_path: Path) -> int:
    epoch = _extract_epoch_from_name(npz_path.name)
    if epoch >= 0:
        return epoch

    epoch_parent = _extract_epoch_from_name(npz_path.parent.name)
    if epoch_parent >= 0:
        return epoch_parent

    sibling_u = npz_path.parent.parent / "u_heatmaps"
    sibling_f = npz_path.parent.parent / "f_heatmaps"
    max_epoch = -1
    for sib_dir, key in ((sibling_u, "u"), (sibling_f, "f")):
        if sib_dir.exists():
            for path in sib_dir.glob(f"{key}_heatmap_*.npz"):
                max_epoch = max(max_epoch, _extract_epoch_from_name(path.name))
    return max_epoch


def infer_sampling_mode_from_path(path: Path) -> str:
    parts = {part.lower() for part in path.parts}
    if "roi_on" in parts:
        return "roi-on"
    if "full_data" in parts:
        return "full-data"
    return "roi-off"


def generate_full_field_np(n_side: int = 201, *, eps: float = 1e-8) -> Tuple[np.ndarray, np.ndarray]:
    xs = np.linspace(-1.0, 1.0, n_side, dtype=np.float64)
    ys = np.linspace(-1.0, 1.0, n_side, dtype=np.float64)
    xg, yg = np.meshgrid(xs, ys, indexing="ij")

    rho = xg * xg + yg * yg
    theta = np.arctan2(xg, yg)
    radius = 0.4 + 0.1 * np.sin(0.0 * theta)
    phi_signed = radius * radius - rho

    u_inside = np.exp(rho)
    a_out = 0.1 * rho * rho - 0.01 * np.log(2.0 * np.sqrt(rho + eps) + eps)
    u_outside = np.sqrt(np.clip(a_out, eps, None))
    u = np.where(phi_signed < 0.0, u_outside, u_inside)

    xy = np.stack([xg.ravel(), yg.ravel()], axis=1)
    return xy.astype(np.float64), u.reshape(-1, 1).astype(np.float64)


def load_uniform_grid_fit_np(
    nx: int,
    ny: int,
    *,
    use_synthetic: bool,
    synthetic_n_side: int,
    ttxt_filename: str,
    circles: Sequence[Tuple[float, float, float]],
    annuli: Sequence[Tuple[float, float, float, float]],
    dense_factor: float,
    drop_boundary: bool,
    xlim: Tuple[float, float],
    ylim: Tuple[float, float],
    tol: float,
    target_total: Optional[int] = None,
) -> Tuple[np.ndarray, np.ndarray]:
    if use_synthetic:
        xy_full, u_full = generate_full_field_np(n_side=synthetic_n_side)
    else:
        xy_full = np.loadtxt(ttxt_filename, usecols=[0, 1], comments="%")
        u_full = np.loadtxt(ttxt_filename, usecols=[2], comments="%").reshape(-1, 1)

    x_vec = np.unique(xy_full[:, 0])
    y_vec = np.unique(xy_full[:, 1])
    tree = cKDTree(xy_full)

    x_tar = np.linspace(x_vec[1], x_vec[-2], nx)
    y_tar = np.linspace(y_vec[1], y_vec[-2], ny)
    xg, yg = np.meshgrid(x_tar, y_tar, indexing="ij")
    ideal_points = np.stack([xg.flatten(), yg.flatten()], axis=1)
    _, idx_basic = tree.query(ideal_points, k=1)
    idx_basic = np.unique(idx_basic)
    idx_dense = np.array([], dtype=int)

    if circles or annuli:
        dense_nx = max(2, int(nx / dense_factor))
        dense_ny = max(2, int(ny / dense_factor))
        x_dense = np.linspace(x_vec[1], x_vec[-2], dense_nx)
        y_dense = np.linspace(y_vec[1], y_vec[-2], dense_ny)
        xd, yd = np.meshgrid(x_dense, y_dense, indexing="ij")
        pts_dense = np.stack([xd.flatten(), yd.flatten()], axis=1)

        for cx, cy, radius in circles:
            mask = ((pts_dense[:, 0] - cx) ** 2 + (pts_dense[:, 1] - cy) ** 2) <= radius**2
            if np.any(mask):
                idx_in = tree.query(pts_dense[mask], k=1)[1]
                idx_dense = np.unique(np.concatenate([idx_dense, idx_in]))

        for cx, cy, r_in, r_out in annuli:
            rin = min(r_in, r_out)
            rout = max(r_in, r_out)
            dist2 = (pts_dense[:, 0] - cx) ** 2 + (pts_dense[:, 1] - cy) ** 2
            mask = (dist2 >= rin**2) & (dist2 <= rout**2)
            if np.any(mask):
                idx_in = tree.query(pts_dense[mask], k=1)[1]
                idx_dense = np.unique(np.concatenate([idx_dense, idx_in]))

    if idx_dense.size > 0:
        idx_dense = np.setdiff1d(idx_dense, idx_basic, assume_unique=False)

    def _filter_boundary(indices: np.ndarray) -> np.ndarray:
        if indices.size == 0:
            return indices
        xy = xy_full[indices]
        mask_inner = (
            (xy[:, 0] > xlim[0] + tol)
            & (xy[:, 0] < xlim[1] - tol)
            & (xy[:, 1] > ylim[0] + tol)
            & (xy[:, 1] < ylim[1] - tol)
        )
        return indices[mask_inner]

    if drop_boundary:
        idx_basic = _filter_boundary(idx_basic)
        idx_dense = _filter_boundary(idx_dense)

    if target_total is not None:
        nb = idx_basic.size
        ne = idx_dense.size
        if nb + ne > target_total:
            rng = np.random.default_rng(1234)
            if nb >= target_total:
                idx_basic = np.array(sorted(rng.choice(idx_basic, size=target_total, replace=False)), dtype=int)
                idx_dense = np.array([], dtype=int)
            else:
                max_dense_keep = max(target_total - nb, 0)
                if ne > max_dense_keep:
                    idx_dense = np.array(sorted(rng.choice(idx_dense, size=max_dense_keep, replace=False)), dtype=int)

    idx_all = np.union1d(idx_basic, idx_dense)
    return xy_full[idx_all].astype(np.float64), u_full[idx_all].astype(np.float64)


def pick_default_model_paths(project_root: Path) -> Dict[str, Path]:
    pinn_final = project_root / "outputs_pinn_single" / "final_fields.npz"
    apinn_final = project_root / "outputs_apinn" / "final_fields.npz"

    pinn_snapshots = list(project_root.glob("outputs_pinn_single/snapshots/epoch_*/fields.npz"))
    apinn_snapshots = list(project_root.glob("outputs_apinn/snapshots/epoch_*/fields.npz"))

    pinn_latest_snap = (
        max(pinn_snapshots, key=lambda p: (_extract_epoch_from_name(p.parent.name), p.stat().st_mtime))
        if pinn_snapshots
        else None
    )
    apinn_latest_snap = (
        max(apinn_snapshots, key=lambda p: (_extract_epoch_from_name(p.parent.name), p.stat().st_mtime))
        if apinn_snapshots
        else None
    )

    pimoe_u_latest = find_latest_heatmap_npz(project_root, "u")
    pimoe_f_latest = find_latest_heatmap_npz(project_root, "f")
    pimoe_fields_fallback = project_root / "outputs_flower" / "roi_off" / "data_output" / "final_fields.npz"

    return {
        "pinn": pinn_final if pinn_final.exists() else (pinn_latest_snap if pinn_latest_snap is not None else Path("")),
        "apinn": apinn_final if apinn_final.exists() else (apinn_latest_snap if apinn_latest_snap is not None else Path("")),
        "pimoe_u": pimoe_u_latest if pimoe_u_latest is not None else pimoe_fields_fallback,
        "pimoe_f": pimoe_f_latest if pimoe_f_latest is not None else pimoe_fields_fallback,
        "pimoe_fields": pimoe_fields_fallback,
    }


def _find_first_key(data: np.lib.npyio.NpzFile, candidates: Sequence[str]) -> Optional[str]:
    for key in candidates:
        if key in data:
            return key
    return None


def _load_axes_from_npz(data: np.lib.npyio.NpzFile, shape: Tuple[int, int]) -> Tuple[np.ndarray, np.ndarray]:
    if "x" in data and "y" in data:
        x_axis = np.asarray(data["x"], dtype=np.float64).reshape(-1)
        y_axis = np.asarray(data["y"], dtype=np.float64).reshape(-1)
        return x_axis, y_axis

    if "Xg" in data and "Yg" in data:
        xg = np.asarray(data["Xg"], dtype=np.float64)
        yg = np.asarray(data["Yg"], dtype=np.float64)
        return xg[:, 0].reshape(-1), yg[0, :].reshape(-1)

    if "bbox" in data:
        bbox = np.asarray(data["bbox"], dtype=np.float64).reshape(-1)
        x_axis = np.linspace(float(bbox[0]), float(bbox[1]), shape[0], dtype=np.float64)
        y_axis = np.linspace(float(bbox[2]), float(bbox[3]), shape[1], dtype=np.float64)
        return x_axis, y_axis

    raise KeyError("Cannot infer grid axes: missing x/y, Xg/Yg, and bbox")


def load_field_record(npz_path: Path, field_name: str) -> Dict[str, np.ndarray | int]:
    with np.load(npz_path) as data:
        true_key = _find_first_key(data, (f"{field_name}_true", f"{field_name}_true_map"))
        pred_key = _find_first_key(data, (f"{field_name}_pred", f"{field_name}_pred_map"))
        if true_key is None or pred_key is None:
            raise KeyError(
                f"{npz_path} missing one of true keys "
                f"('{field_name}_true', '{field_name}_true_map') or pred keys "
                f"('{field_name}_pred', '{field_name}_pred_map')"
            )

        true_map = np.asarray(data[true_key], dtype=np.float64)
        pred_map = np.asarray(data[pred_key], dtype=np.float64)
        x_axis, y_axis = _load_axes_from_npz(data, true_map.shape)

        epoch = -1
        if "epoch" in data:
            epoch_values = np.asarray(data["epoch"]).reshape(-1)
            if epoch_values.size > 0:
                epoch = int(epoch_values[0])
        if epoch < 0:
            epoch = infer_epoch_from_npz_path(npz_path)

    return {
        "true_map": true_map,
        "pred_map": pred_map,
        "x_axis": x_axis,
        "y_axis": y_axis,
        "epoch": epoch,
    }


def l2_relative(true_values: np.ndarray, pred_values: np.ndarray) -> float:
    diff = (pred_values - true_values).reshape(-1)
    true_flat = true_values.reshape(-1)
    return float(np.linalg.norm(diff, ord=2) / (np.linalg.norm(true_flat, ord=2) + 1e-15))


def format_display(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}E}"


def format_percent(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}f}%"


def _nearest_indices(axis: np.ndarray, values: np.ndarray) -> np.ndarray:
    axis = np.asarray(axis, dtype=np.float64).reshape(-1)
    values = np.asarray(values, dtype=np.float64).reshape(-1)

    idx = np.searchsorted(axis, values)
    idx = np.clip(idx, 0, axis.size - 1)
    left = np.clip(idx - 1, 0, axis.size - 1)
    choose_left = np.abs(values - axis[left]) <= np.abs(values - axis[idx])
    idx[choose_left] = left[choose_left]
    return idx


def sample_grid_at_points(
    field_map: np.ndarray,
    x_axis: np.ndarray,
    y_axis: np.ndarray,
    xy_points: np.ndarray,
) -> np.ndarray:
    ix = _nearest_indices(x_axis, xy_points[:, 0])
    iy = _nearest_indices(y_axis, xy_points[:, 1])
    return np.asarray(field_map, dtype=np.float64)[ix, iy].reshape(-1, 1)


def load_label_dataset(
    project_root: Path,
    *,
    sampling_mode: Optional[str],
) -> Tuple[np.ndarray, np.ndarray]:
    data_cfg = DataConfig()
    data_path = (project_root / data_cfg.ttxt_filename).resolve()

    if sampling_mode is None:
        xy_fit, u_fit = load_uniform_grid_fit_np(
            nx=data_cfg.nx,
            ny=data_cfg.ny,
            use_synthetic=data_cfg.use_synthetic,
            synthetic_n_side=data_cfg.synthetic_n_side,
            ttxt_filename=str(data_path),
            circles=list(data_cfg.circles),
            annuli=list(data_cfg.annuli),
            dense_factor=data_cfg.dense_factor,
            drop_boundary=data_cfg.drop_boundary,
            xlim=data_cfg.xlim,
            ylim=data_cfg.ylim,
            tol=data_cfg.tol,
        )
    else:
        sampling_cfg = SAMPLING_CONFIGS[sampling_mode]
        xy_fit, u_fit = load_uniform_grid_fit_np(
            nx=sampling_cfg["nx"],
            ny=sampling_cfg["ny"],
            use_synthetic=data_cfg.use_synthetic,
            synthetic_n_side=data_cfg.synthetic_n_side,
            ttxt_filename=str(data_path),
            circles=list(data_cfg.circles),
            annuli=list(data_cfg.annuli),
            dense_factor=sampling_cfg["dense_factor"],
            drop_boundary=sampling_cfg["drop_boundary"],
            xlim=sampling_cfg["xlim"],
            ylim=sampling_cfg["ylim"],
            tol=sampling_cfg["tol"],
            target_total=sampling_cfg.get("target_total"),
        )

    return xy_fit.astype(np.float64), u_fit.astype(np.float64)


def pointwise_relative_percentages(
    true_map: np.ndarray,
    pred_map: np.ndarray,
    thresholds: Sequence[float],
    *,
    eps: float = 1e-12,
) -> Dict[float, float]:
    rel_map = np.abs(np.asarray(pred_map, dtype=np.float64) - np.asarray(true_map, dtype=np.float64))
    rel_map = rel_map / np.maximum(np.abs(np.asarray(true_map, dtype=np.float64)), eps)
    return {thr: float(np.mean(rel_map <= thr) * 100.0) for thr in thresholds}


def threshold_column_name(threshold: float) -> str:
    pct = threshold * 100.0
    if abs(pct - round(pct)) < 1e-12:
        pct_str = str(int(round(pct)))
    else:
        pct_str = f"{pct:.3f}".rstrip("0").rstrip(".")
    return f"points_within_{pct_str}pct"


def threshold_column_label(threshold: float) -> str:
    pct = threshold * 100.0
    if abs(pct - round(pct)) < 1e-12:
        pct_str = str(int(round(pct)))
    else:
        pct_str = f"{pct:.3f}".rstrip("0").rstrip(".")
    return f"Points within {pct_str}% Error (%)"


def threshold_header(threshold: float) -> str:
    pct = threshold * 100.0
    if abs(pct - round(pct)) < 1e-12:
        pct_str = str(int(round(pct)))
    else:
        pct_str = f"{pct:.3f}".rstrip("0").rstrip(".")
    return f"{pct_str}%"


def relabel_rows(
    rows: List[Dict[str, object]],
    fieldnames: List[str],
    label_map: Dict[str, str],
) -> Tuple[List[Dict[str, object]], List[str]]:
    labeled_fieldnames = [label_map.get(field, field) for field in fieldnames]
    labeled_rows: List[Dict[str, object]] = []
    for row in rows:
        labeled_rows.append({label_map.get(field, field): row[field] for field in fieldnames})
    return labeled_rows, labeled_fieldnames


def write_excel_sheets(
    output_path: Path,
    tables: Sequence[Tuple[str, List[Dict[str, object]], List[str]]],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows, fieldnames in tables:
            frame = pd.DataFrame(rows, columns=fieldnames)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(output_path)
    text_headers = {"Model", "Label Sampling"}
    epoch_headers = {"Epoch", "Epoch (u)", "Epoch (f)"}
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header in text_headers:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                if header in epoch_headers:
                    if isinstance(cell.value, str):
                        try:
                            cell.value = int(float(cell.value))
                        except ValueError:
                            continue
                    cell.number_format = "0"
        for col_idx, _ in enumerate(headers, start=1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                text = "" if value is None else str(value)
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    wb.save(output_path)


def print_table(title: str, rows: List[Dict[str, object]], fieldnames: List[str]) -> None:
    print(title)
    if not rows:
        print("(empty)")
        print("")
        return

    widths: Dict[str, int] = {}
    for field in fieldnames:
        widths[field] = len(field)
        for row in rows:
            widths[field] = max(widths[field], len(str(row[field])))

    header = " | ".join(f"{field:{widths[field]}}" for field in fieldnames)
    divider = "-+-".join("-" * widths[field] for field in fieldnames)
    print(header)
    print(divider)
    for row in rows:
        print(" | ".join(f"{str(row[field]):{widths[field]}}" for field in fieldnames))
    print("")


def main() -> None:
    project_root = Path(__file__).resolve().parent
    defaults = pick_default_model_paths(project_root)

    parser = argparse.ArgumentParser(
        description="Export two tables: model relative errors and PDE pointwise relative-error percentages."
    )
    parser.add_argument("--pinn", type=str, default=str(defaults["pinn"]))
    parser.add_argument("--apinn", type=str, default=str(defaults["apinn"]))
    parser.add_argument("--pimoe-u", type=str, default=str(defaults["pimoe_u"]))
    parser.add_argument("--pimoe-f", "--pimoe", dest="pimoe_f", type=str, default=str(defaults["pimoe_f"]))
    parser.add_argument("--pimoe-fields", type=str, default=str(defaults["pimoe_fields"]))
    parser.add_argument(
        "--excel",
        type=str,
        default=str(project_root / "error_summary_tables.xlsx"),
        help="Output Excel workbook path. Different tables are written to different sheets.",
    )
    parser.add_argument(
        "--thresholds",
        type=str,
        default="0.001,0.002,0.005,0.01",
        help="Comma-separated relative-error thresholds, e.g. 0.01,0.02,0.05",
    )
    args = parser.parse_args()

    pinn_path = Path(args.pinn).expanduser().resolve()
    apinn_path = Path(args.apinn).expanduser().resolve()
    pimoe_u_path = Path(args.pimoe_u).expanduser().resolve()
    pimoe_f_path = Path(args.pimoe_f).expanduser().resolve()
    pimoe_fields_path = Path(args.pimoe_fields).expanduser().resolve()
    excel_path = Path(args.excel).expanduser().resolve()
    if excel_path.suffix.lower() != ".xlsx":
        excel_path = excel_path.with_suffix(".xlsx")
    thresholds = [float(item.strip()) for item in args.thresholds.split(",") if item.strip()]

    for label, path in (
        ("PINN", pinn_path),
        ("APINN", apinn_path),
        ("ADD-PINNs u", pimoe_u_path),
        ("ADD-PINNs f", pimoe_f_path),
        ("ADD-PINNs fields", pimoe_fields_path),
    ):
        if not path.exists():
            raise FileNotFoundError(f"{label} file not found: {path}")

    pinn_u = load_field_record(pinn_path, "u")
    pinn_f = load_field_record(pinn_path, "f")
    apinn_u = load_field_record(apinn_path, "u")
    apinn_f = load_field_record(apinn_path, "f")
    pimoe_u = load_field_record(pimoe_u_path, "u")
    pimoe_f = load_field_record(pimoe_f_path, "f")
    pimoe_fields_u = load_field_record(pimoe_fields_path, "u")

    xy_label_pinn, u_label_pinn = load_label_dataset(project_root, sampling_mode=None)
    xy_label_apinn, u_label_apinn = load_label_dataset(project_root, sampling_mode=None)
    pimoe_sampling_mode = infer_sampling_mode_from_path(pimoe_fields_path)
    xy_label_pimoe, u_label_pimoe = load_label_dataset(project_root, sampling_mode=pimoe_sampling_mode)

    label_pred_pinn = sample_grid_at_points(pinn_u["pred_map"], pinn_u["x_axis"], pinn_u["y_axis"], xy_label_pinn)
    label_pred_apinn = sample_grid_at_points(apinn_u["pred_map"], apinn_u["x_axis"], apinn_u["y_axis"], xy_label_apinn)
    label_pred_pimoe = sample_grid_at_points(
        pimoe_fields_u["pred_map"],
        pimoe_fields_u["x_axis"],
        pimoe_fields_u["y_axis"],
        xy_label_pimoe,
    )

    table1_fieldnames = ["Model", "L2 Error on Label Data"] + [threshold_header(thr) for thr in thresholds]
    table2_fieldnames = ["Model"] + [threshold_header(thr) for thr in thresholds]

    table1_rows: List[Dict[str, object]] = []
    for model_name, label_true, label_pred, u_record in (
        ("PINN", u_label_pinn, label_pred_pinn, pinn_u),
        ("APINN", u_label_apinn, label_pred_apinn, apinn_u),
        ("ADD-PINNs", u_label_pimoe, label_pred_pimoe, pimoe_u),
    ):
        row = {
            "Model": model_name,
            "L2 Error on Label Data": format_display(l2_relative(label_true, label_pred)),
        }
        percentages = pointwise_relative_percentages(u_record["true_map"], u_record["pred_map"], thresholds)
        for threshold in thresholds:
            row[threshold_header(threshold)] = format_percent(percentages[threshold])
        table1_rows.append(row)

    table2_rows: List[Dict[str, object]] = []
    for model_name, record in (
        ("PINN", pinn_f),
        ("APINN", apinn_f),
        ("ADD-PINNs", pimoe_f),
    ):
        row = {"Model": model_name}
        percentages = pointwise_relative_percentages(record["true_map"], record["pred_map"], thresholds)
        for threshold in thresholds:
            row[threshold_header(threshold)] = format_percent(percentages[threshold])
        table2_rows.append(row)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel_sheets(
        excel_path,
        [
            ("Table1", table1_rows, table1_fieldnames),
            ("Table2", table2_rows, table2_fieldnames),
        ],
    )

    print_table("Table 1: label-data L2 error and full-field u pointwise relative-error percentages", table1_rows, table1_fieldnames)
    print_table("Table 2: full-field PDE pointwise relative-error percentages", table2_rows, table2_fieldnames)
    print(f"Saved: {excel_path}")


if __name__ == "__main__":
    main()
