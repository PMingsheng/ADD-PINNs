#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.interpolate import RegularGridInterpolator

from config import DataConfig, SAMPLING_CONFIGS
from data import load_uniform_grid_fit


def load_reference(txt_filename: Path) -> Tuple[np.ndarray, np.ndarray]:
    xy = np.loadtxt(txt_filename, usecols=[0, 1], comments="%").astype(np.float64)
    t = np.loadtxt(txt_filename, usecols=[2], comments="%").reshape(-1, 1).astype(np.float64)
    return xy, t


def infer_sampling_mode_from_path(path: Path) -> str:
    parts = {part.lower() for part in path.parts}
    if "output_roi_on" in parts or "roi_on" in parts:
        return "roi-on"
    if "output_full_data" in parts or "full_data" in parts:
        return "full-data"
    return "roi-off"


def _extract_epoch_from_name(name: str) -> int:
    digits = "".join(ch for ch in name if ch.isdigit())
    return int(digits) if digits else -1


def find_latest_phi_snapshot(snapshot_dir: Path) -> Optional[Path]:
    candidates = list(snapshot_dir.glob("phi_epoch_*.npz"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: (_extract_epoch_from_name(p.name), p.stat().st_mtime))


def _load_axes_and_fields(npz_path: Path) -> Dict[str, np.ndarray]:
    with np.load(npz_path) as zf:
        data = {k: zf[k] for k in zf.files}

    if "x" not in data or "y" not in data:
        raise KeyError(f"{npz_path} missing x/y axes.")

    t_key = "T_pred" if "T_pred" in data else "T"
    if t_key not in data:
        raise KeyError(f"{npz_path} missing T/T_pred.")

    epoch = -1
    if "epoch" in data:
        epoch_arr = np.asarray(data["epoch"]).reshape(-1)
        if epoch_arr.size > 0:
            epoch = int(epoch_arr[0])
    if epoch < 0:
        epoch = _extract_epoch_from_name(npz_path.name)

    f1 = float(np.asarray(data["f1"]).reshape(-1)[0]) if "f1" in data else float("nan")
    f2 = float(np.asarray(data["f2"]).reshape(-1)[0]) if "f2" in data else float("nan")

    return {
        "x": np.asarray(data["x"], dtype=np.float64).reshape(-1),
        "y": np.asarray(data["y"], dtype=np.float64).reshape(-1),
        "T": np.asarray(data[t_key], dtype=np.float64),
        "epoch": np.asarray([epoch], dtype=np.int64),
        "f1": np.asarray([f1], dtype=np.float64),
        "f2": np.asarray([f2], dtype=np.float64),
    }


def load_pde_residual_record(npz_path: Path) -> Dict[str, object]:
    with np.load(npz_path) as zf:
        data = {k: zf[k] for k in zf.files}

    if "pde_residual" not in data:
        raise KeyError(f"{npz_path} missing pde_residual.")
    if "x" not in data or "y" not in data:
        raise KeyError(f"{npz_path} missing x/y axes.")

    epoch = -1
    if "epoch" in data:
        epoch_arr = np.asarray(data["epoch"]).reshape(-1)
        if epoch_arr.size > 0:
            epoch = int(epoch_arr[0])
    if epoch < 0:
        epoch = _extract_epoch_from_name(npz_path.name)

    return {
        "epoch": epoch,
        "x": np.asarray(data["x"], dtype=np.float64).reshape(-1),
        "y": np.asarray(data["y"], dtype=np.float64).reshape(-1),
        "pde_residual": np.abs(np.asarray(data["pde_residual"], dtype=np.float64)),
    }


def sample_prediction_at_points(
    field_map: np.ndarray,
    x_axis: np.ndarray,
    y_axis: np.ndarray,
    xy_points: np.ndarray,
) -> np.ndarray:
    interp = RegularGridInterpolator(
        (np.asarray(x_axis, dtype=np.float64), np.asarray(y_axis, dtype=np.float64)),
        np.asarray(field_map, dtype=np.float64),
        method="linear",
        bounds_error=False,
        fill_value=np.nan,
    )
    return np.asarray(interp(np.asarray(xy_points, dtype=np.float64)), dtype=np.float64).reshape(-1, 1)


def relative_l2(true_values: np.ndarray, pred_values: np.ndarray) -> float:
    diff = np.asarray(pred_values, dtype=np.float64).reshape(-1) - np.asarray(true_values, dtype=np.float64).reshape(-1)
    denom = np.linalg.norm(np.asarray(true_values, dtype=np.float64).reshape(-1), ord=2) + 1e-15
    return float(np.linalg.norm(diff, ord=2) / denom)


def pointwise_relative_percentages(
    true_values: np.ndarray,
    pred_values: np.ndarray,
    thresholds: Sequence[float],
    *,
    eps: float = 1e-12,
) -> Dict[float, float]:
    true_arr = np.asarray(true_values, dtype=np.float64).reshape(-1)
    pred_arr = np.asarray(pred_values, dtype=np.float64).reshape(-1)
    rel = np.abs(pred_arr - true_arr) / np.maximum(np.abs(true_arr), eps)
    return {thr: float(np.mean(rel <= thr) * 100.0) for thr in thresholds}


def residual_within_threshold_percentages(
    residual_values: np.ndarray,
    thresholds: Sequence[float],
) -> Dict[float, float]:
    residual = np.abs(np.asarray(residual_values, dtype=np.float64)).reshape(-1)
    return {thr: float(np.mean(residual <= thr) * 100.0) for thr in thresholds}


def build_true_source_field(
    x_axis: np.ndarray,
    y_axis: np.ndarray,
    *,
    f1_true: float,
    f2_true: float,
) -> np.ndarray:
    data_cfg = DataConfig()
    if not data_cfg.circles:
        raise ValueError("DataConfig.circles is empty.")
    cx, cy, radius = (float(v) for v in data_cfg.circles[0])
    xx, yy = np.meshgrid(
        np.asarray(x_axis, dtype=np.float64),
        np.asarray(y_axis, dtype=np.float64),
        indexing="ij",
    )
    phi = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2 + 1e-12) - radius
    return np.where(phi >= 0.0, float(f1_true), float(f2_true))


def normalized_residual_within_threshold_percentages(
    residual_values: np.ndarray,
    reference_scale: np.ndarray,
    thresholds: Sequence[float],
    *,
    eps: float = 1e-12,
) -> Dict[float, float]:
    residual = np.abs(np.asarray(residual_values, dtype=np.float64))
    scale = np.maximum(np.abs(np.asarray(reference_scale, dtype=np.float64)), eps)
    normalized = (residual / scale).reshape(-1)
    return {thr: float(np.mean(normalized <= thr) * 100.0) for thr in thresholds}


def format_display(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}E}"


def format_percent(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}f}%"


def threshold_header(threshold: float) -> str:
    pct = threshold * 100.0
    if abs(pct - round(pct)) < 1e-12:
        pct_str = str(int(round(pct)))
    else:
        pct_str = f"{pct:.3f}".rstrip("0").rstrip(".")
    return f"{pct_str}%"


def tau_header(threshold: float) -> str:
    return f"{threshold:.0e}".replace("e-0", "e-").replace("e+0", "e+")


def write_excel_sheets(
    output_path: Path,
    tables: Sequence[Tuple[str, List[Dict[str, object]], List[str]]],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows, fieldnames in tables:
            frame = pd.DataFrame(rows, columns=fieldnames)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    text_headers = {"Model"}
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        for col_idx, _header in enumerate(headers, start=1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                text = "" if value is None else str(value)
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        for col_idx, header in enumerate(headers, start=1):
            if header in text_headers:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
    wb.save(output_path)


def print_table(title: str, rows: List[Dict[str, object]], fieldnames: List[str]) -> None:
    print(title)
    if not rows:
        print("(empty)")
        print("")
        return

    widths: Dict[str, int] = {}
    for field in fieldnames:
        widths[field] = len(field)
        for row in rows:
            widths[field] = max(widths[field], len(str(row[field])))

    header = " | ".join(f"{field:{widths[field]}}" for field in fieldnames)
    divider = "-+-".join("-" * widths[field] for field in fieldnames)
    print(header)
    print(divider)
    for row in rows:
        print(" | ".join(f"{str(row[field]):{widths[field]}}" for field in fieldnames))
    print("")


def load_label_dataset(project_root: Path, *, sampling_mode: str) -> Tuple[np.ndarray, np.ndarray]:
    data_cfg = DataConfig()
    if sampling_mode not in SAMPLING_CONFIGS:
        raise ValueError(f"Unknown sampling mode: {sampling_mode}")
    cfg = SAMPLING_CONFIGS[sampling_mode]
    xy_fit, t_fit = load_uniform_grid_fit(
        nx=cfg["nx"],
        ny=cfg["ny"],
        ttxt_filename=str((project_root / cfg["ttxt_filename"]).resolve()),
        device="cpu",
        circles=data_cfg.circles,
        dense_factor=cfg["dense_factor"],
        drop_boundary=cfg["drop_boundary"],
        xlim=cfg["xlim"],
        ylim=cfg["ylim"],
        tol=cfg["tol"],
        target_total=cfg.get("target_total"),
    )
    return xy_fit.detach().cpu().numpy().astype(np.float64), t_fit.detach().cpu().numpy().astype(np.float64)


def evaluate_model(npz_path: Path, xy_ref: np.ndarray, t_ref: np.ndarray) -> Dict[str, object]:
    fields = _load_axes_and_fields(npz_path)
    t_pred = sample_prediction_at_points(fields["T"], fields["x"], fields["y"], xy_ref)
    return {
        "epoch": int(fields["epoch"][0]),
        "f1": float(fields["f1"][0]),
        "f2": float(fields["f2"][0]),
        "t_relative_l2": relative_l2(t_ref, t_pred),
        "t_pred_points": t_pred,
    }


def main() -> None:
    project_root = Path(__file__).resolve().parent
    default_pimoe_final = project_root / "output_roi_off" / "data_output" / "final_fields.npz"
    default_pimoe_pde = find_latest_phi_snapshot(project_root / "output_roi_off" / "phi_snapshots")
    parser = argparse.ArgumentParser(
        description="Export two Possion tables in one workbook."
    )
    parser.add_argument("--txt", type=str, default=str(project_root / "Possion.txt"))
    parser.add_argument("--pimoe", type=str, default=str(default_pimoe_final))
    parser.add_argument("--apinn", type=str, default=str(project_root / "outputs_apinn" / "final_fields.npz"))
    parser.add_argument("--pinn", type=str, default=str(project_root / "outputs_pinn_single" / "final_fields.npz"))
    parser.add_argument(
        "--pimoe-pde",
        type=str,
        default=str(default_pimoe_pde) if default_pimoe_pde is not None else "",
        help="ADD-PINNs PDE residual snapshot (.npz). Defaults to the latest phi snapshot in output_roi_off.",
    )
    parser.add_argument("--csv", type=str, default=str(project_root / "error_summary_t_l2.csv"))
    parser.add_argument(
        "--excel",
        type=str,
        default=str(project_root / "error_summary_tables.xlsx"),
        help="Output Excel workbook path. Different tables are written to different sheets.",
    )
    parser.add_argument(
        "--thresholds",
        type=str,
        default="0.001,0.002,0.005,0.01",
        help="Comma-separated relative-error thresholds for T, e.g. 0.001,0.002,0.005,0.01",
    )
    parser.add_argument(
        "--pde-thresholds",
        type=str,
        default="1e-2,2e-2,5e-2,1e-1,2e-1,5e-1",
        help="Comma-separated absolute thresholds tau for |pde_residual|.",
    )
    parser.add_argument(
        "--pde-relative-thresholds",
        type=str,
        default="0.01,0.02,0.05,0.10",
        help="Comma-separated thresholds tau for |pde_residual| / |f_true|.",
    )
    parser.add_argument(
        "--f1-true",
        type=float,
        default=10.0,
        help="True source value used where phi >= 0 (outside the circle).",
    )
    parser.add_argument(
        "--f2-true",
        type=float,
        default=5.0,
        help="True source value used where phi < 0 (inside the circle).",
    )
    args = parser.parse_args()

    txt_path = Path(args.txt).expanduser().resolve()
    pinn_path = Path(args.pinn).expanduser().resolve()
    apinn_path = Path(args.apinn).expanduser().resolve()
    pimoe_path = Path(args.pimoe).expanduser().resolve()
    if not args.pimoe_pde.strip():
        raise FileNotFoundError("ADD-PINNs PDE snapshot not found. Set --pimoe-pde explicitly.")
    pimoe_pde_path = Path(args.pimoe_pde).expanduser().resolve()
    csv_path = Path(args.csv).expanduser().resolve()
    excel_path = Path(args.excel).expanduser().resolve()
    if excel_path.suffix.lower() != ".xlsx":
        excel_path = excel_path.with_suffix(".xlsx")
    thresholds = [float(item.strip()) for item in args.thresholds.split(",") if item.strip()]
    pde_thresholds = [float(item.strip()) for item in args.pde_thresholds.split(",") if item.strip()]
    pde_relative_thresholds = [float(item.strip()) for item in args.pde_relative_thresholds.split(",") if item.strip()]

    for label, path in (
        ("PINN", pinn_path),
        ("APINN", apinn_path),
        ("ADD-PINNs", pimoe_path),
        ("ADD-PINNs PDE", pimoe_pde_path),
        ("Reference", txt_path),
    ):
        if not path.exists():
            raise FileNotFoundError(f"{label} file not found: {path}")

    xy_ref, t_ref = load_reference(txt_path)

    pinn_eval = evaluate_model(pinn_path, xy_ref, t_ref)
    apinn_eval = evaluate_model(apinn_path, xy_ref, t_ref)
    pimoe_eval = evaluate_model(pimoe_path, xy_ref, t_ref)

    pinn_pde = load_pde_residual_record(pinn_path)
    apinn_pde = load_pde_residual_record(apinn_path)
    pimoe_pde = load_pde_residual_record(pimoe_pde_path)

    pinn_sampling_mode = "roi-off"
    apinn_sampling_mode = "roi-off"
    pimoe_sampling_mode = infer_sampling_mode_from_path(pimoe_path)

    xy_label_pinn, t_label_pinn = load_label_dataset(project_root, sampling_mode=pinn_sampling_mode)
    xy_label_apinn, t_label_apinn = load_label_dataset(project_root, sampling_mode=apinn_sampling_mode)
    xy_label_pimoe, t_label_pimoe = load_label_dataset(project_root, sampling_mode=pimoe_sampling_mode)

    pinn_fields = _load_axes_and_fields(pinn_path)
    apinn_fields = _load_axes_and_fields(apinn_path)
    pimoe_fields = _load_axes_and_fields(pimoe_path)

    label_pred_pinn = sample_prediction_at_points(pinn_fields["T"], pinn_fields["x"], pinn_fields["y"], xy_label_pinn)
    label_pred_apinn = sample_prediction_at_points(apinn_fields["T"], apinn_fields["x"], apinn_fields["y"], xy_label_apinn)
    label_pred_pimoe = sample_prediction_at_points(pimoe_fields["T"], pimoe_fields["x"], pimoe_fields["y"], xy_label_pimoe)

    table1_fieldnames = ["Model", "L2 Error on Label Data"] + [threshold_header(thr) for thr in thresholds]
    table2_fieldnames = ["Model"] + [tau_header(thr) for thr in pde_thresholds]
    table3_fieldnames = ["Model"] + [threshold_header(thr) for thr in pde_relative_thresholds]

    table1_rows: List[Dict[str, object]] = []
    for model_name, label_true, label_pred, pred_values in (
        ("PINN", t_label_pinn, label_pred_pinn, pinn_eval["t_pred_points"]),
        ("APINN", t_label_apinn, label_pred_apinn, apinn_eval["t_pred_points"]),
        ("ADD-PINNs", t_label_pimoe, label_pred_pimoe, pimoe_eval["t_pred_points"]),
    ):
        row = {
            "Model": model_name,
            "L2 Error on Label Data": format_display(relative_l2(label_true, label_pred)),
        }
        percentages = pointwise_relative_percentages(t_ref, pred_values, thresholds)
        for threshold in thresholds:
            row[threshold_header(threshold)] = format_percent(percentages[threshold])
        table1_rows.append(row)

    table2_rows: List[Dict[str, object]] = []
    for model_name, record in (
        ("PINN", pinn_pde),
        ("APINN", apinn_pde),
        ("ADD-PINNs", pimoe_pde),
    ):
        row = {"Model": model_name}
        percentages = residual_within_threshold_percentages(record["pde_residual"], pde_thresholds)
        for threshold in pde_thresholds:
            row[tau_header(threshold)] = format_percent(percentages[threshold])
        table2_rows.append(row)

    table3_rows: List[Dict[str, object]] = []
    for model_name, record in (
        ("PINN", pinn_pde),
        ("APINN", apinn_pde),
        ("ADD-PINNs", pimoe_pde),
    ):
        f_true = build_true_source_field(
            record["x"],
            record["y"],
            f1_true=args.f1_true,
            f2_true=args.f2_true,
        )
        row = {"Model": model_name}
        percentages = normalized_residual_within_threshold_percentages(
            record["pde_residual"],
            f_true,
            pde_relative_thresholds,
        )
        for threshold in pde_relative_thresholds:
            row[threshold_header(threshold)] = format_percent(percentages[threshold])
        table3_rows.append(row)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel_sheets(
        excel_path,
        [
            ("Table1", table1_rows, table1_fieldnames),
            ("Table2", table2_rows, table2_fieldnames),
            ("Table3", table3_rows, table3_fieldnames),
        ],
    )

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=table1_fieldnames)
        writer.writeheader()
        writer.writerows(table1_rows)

    print_table(
        "Table 1: label-data L2 error and full-field T pointwise relative-error percentages",
        table1_rows,
        table1_fieldnames,
    )
    print_table(
        "Table 2: PDE residual percentages with |pde_residual| <= tau",
        table2_rows,
        table2_fieldnames,
    )
    print_table(
        "Table 3: normalized PDE residual percentages with |pde_residual| / |f_true| <= tau",
        table3_rows,
        table3_fieldnames,
    )
    print(f"Saved: {csv_path}")
    print(f"Saved: {excel_path}")


if __name__ == "__main__":
    main()
