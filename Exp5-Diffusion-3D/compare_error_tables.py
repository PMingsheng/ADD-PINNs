#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MODEL_ORDER = ["PINN", "APINN", "ADD-PINNs"]

DEFAULT_MODEL_PATHS = {
    "PINN": (
        "outputs_pinn3d_c1_sphere/data_output/final_fields.npz",
        "outputs_pinn3d_c1_sphere/final_fields.npz",
    ),
    "APINN": (
        "outputs_apinn3d_c1_sphere/data_output/final_fields.npz",
        "outputs_apinn3d_c1_sphere/final_fields.npz",
    ),
    "ADD-PINNs": (
        "outputs_add_pinns3d_c1_sphere/data_output/final_fields.npz",
        "outputs_add_pinns3d_c1_sphere/final_fields.npz",
    ),
}


def _resolve_existing_path(project_root: Path, rel_paths: Sequence[str]) -> Path:
    tried: List[str] = []
    for rel_path in rel_paths:
        path = (project_root / rel_path).resolve()
        tried.append(str(path))
        if path.exists():
            return path
    raise FileNotFoundError("NPZ file not found. Tried:\n" + "\n".join(tried))


def _extract_scalar(data: Dict[str, np.ndarray], key: str) -> float:
    if key not in data:
        raise KeyError(f"Missing key: {key}")
    arr = np.asarray(data[key], dtype=np.float64).reshape(-1)
    if arr.size == 0:
        raise ValueError(f"Empty scalar field: {key}")
    return float(arr[0])


def _load_npz(npz_path: Path) -> Dict[str, np.ndarray]:
    with np.load(npz_path) as zf:
        return {k: zf[k] for k in zf.files}


def pointwise_relative_percentages(
    true_values: np.ndarray,
    pred_values: np.ndarray,
    thresholds: Sequence[float],
    *,
    eps: float = 1e-12,
) -> Dict[float, float]:
    true_arr = np.asarray(true_values, dtype=np.float64).reshape(-1)
    pred_arr = np.asarray(pred_values, dtype=np.float64).reshape(-1)
    rel = np.abs(pred_arr - true_arr) / np.maximum(np.abs(true_arr), eps)
    return {thr: float(np.mean(rel <= thr) * 100.0) for thr in thresholds}


def residual_within_threshold_percentages(
    residual_values: np.ndarray,
    thresholds: Sequence[float],
) -> Dict[float, float]:
    residual = np.abs(np.asarray(residual_values, dtype=np.float64)).reshape(-1)
    return {thr: float(np.mean(residual <= thr) * 100.0) for thr in thresholds}


def normalized_residual_within_threshold_percentages(
    residual_values: np.ndarray,
    reference_scale: np.ndarray,
    thresholds: Sequence[float],
    *,
    eps: float = 1e-12,
) -> Dict[float, float]:
    residual = np.abs(np.asarray(residual_values, dtype=np.float64))
    scale = np.maximum(np.abs(np.asarray(reference_scale, dtype=np.float64)), eps)
    normalized = (residual / scale).reshape(-1)
    return {thr: float(np.mean(normalized <= thr) * 100.0) for thr in thresholds}


def format_display(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}E}"


def format_percent(value: float, digits: int = 2) -> str:
    value = float(value)
    if not np.isfinite(value):
        return str(value)
    return f"{value:.{digits}f}%"


def threshold_header(threshold: float) -> str:
    pct = threshold * 100.0
    if abs(pct - round(pct)) < 1e-12:
        pct_str = str(int(round(pct)))
    else:
        pct_str = f"{pct:.4f}".rstrip("0").rstrip(".")
    return f"{pct_str}%"


def tau_header(threshold: float) -> str:
    if threshold >= 1.0 and abs(threshold - round(threshold)) < 1e-12:
        return str(int(round(threshold)))
    return f"{threshold:.0e}".replace("e-0", "e-").replace("e+0", "e+")


def write_excel_sheets(
    output_path: Path,
    tables: Sequence[Tuple[str, List[Dict[str, object]], List[str]]],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows, fieldnames in tables:
            frame = pd.DataFrame(rows, columns=fieldnames)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    text_headers = {"Model"}
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        for col_idx, _header in enumerate(headers, start=1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                text = "" if value is None else str(value)
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        for col_idx, header in enumerate(headers, start=1):
            if header in text_headers:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
    wb.save(output_path)


def print_table(title: str, rows: List[Dict[str, object]], fieldnames: List[str]) -> None:
    print(title)
    if not rows:
        print("(empty)")
        print("")
        return

    widths: Dict[str, int] = {}
    for field in fieldnames:
        widths[field] = len(field)
        for row in rows:
            widths[field] = max(widths[field], len(str(row[field])))

    header = " | ".join(f"{field:{widths[field]}}" for field in fieldnames)
    divider = "-+-".join("-" * widths[field] for field in fieldnames)
    print(header)
    print(divider)
    for row in rows:
        print(" | ".join(f"{str(row[field]):{widths[field]}}" for field in fieldnames))
    print("")


def parse_args() -> argparse.Namespace:
    project_root = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Export 3D error summary tables for ADD-PINNs, PINN, and APINN."
    )
    parser.add_argument("--root", type=str, default=str(project_root))
    parser.add_argument("--pimoe", type=str, default=None)
    parser.add_argument("--pinn", type=str, default=None)
    parser.add_argument("--apinn", type=str, default=None)
    parser.add_argument(
        "--u-thresholds",
        type=str,
        default="1e-4,2e-4,5e-4,1e-3",
        help="Pointwise relative-error thresholds for u.",
    )
    parser.add_argument(
        "--f-thresholds",
        type=str,
        default="1e-1,2e-1,5e-1,1,2,5",
        help="Absolute thresholds tau for |f_residual|.",
    )
    parser.add_argument(
        "--f-relative-thresholds",
        type=str,
        default="0.01,0.02,0.05,0.10",
        help="Thresholds tau for |f_residual| / |f_true|.",
    )
    parser.add_argument(
        "--csv",
        type=str,
        default=str(project_root / "error_summary_u_l2.csv"),
        help="CSV output for Table1.",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default=str(project_root / "error_summary_tables.xlsx"),
        help="Excel workbook output path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    project_root = Path(args.root).expanduser().resolve()

    model_paths: Dict[str, Path] = {}
    overrides = {
        "PINN": args.pinn,
        "APINN": args.apinn,
        "ADD-PINNs": args.pimoe,
    }
    for model_name in MODEL_ORDER:
        if overrides[model_name]:
            model_paths[model_name] = Path(overrides[model_name]).expanduser().resolve()
        else:
            model_paths[model_name] = _resolve_existing_path(project_root, DEFAULT_MODEL_PATHS[model_name])

    u_thresholds = [float(item.strip()) for item in args.u_thresholds.split(",") if item.strip()]
    f_thresholds = [float(item.strip()) for item in args.f_thresholds.split(",") if item.strip()]
    f_relative_thresholds = [float(item.strip()) for item in args.f_relative_thresholds.split(",") if item.strip()]

    model_data = {model_name: _load_npz(model_paths[model_name]) for model_name in MODEL_ORDER}

    table1_fieldnames = ["Model", "L2 Error on Label Data"] + [threshold_header(thr) for thr in u_thresholds]
    table2_fieldnames = ["Model"] + [tau_header(thr) for thr in f_thresholds]
    table3_fieldnames = ["Model"] + [threshold_header(thr) for thr in f_relative_thresholds]

    table1_rows: List[Dict[str, object]] = []
    table2_rows: List[Dict[str, object]] = []
    table3_rows: List[Dict[str, object]] = []

    for model_name in MODEL_ORDER:
        data = model_data[model_name]

        u_true = np.asarray(data["u_true"], dtype=np.float64)
        u_pred = np.asarray(data["u_pred"], dtype=np.float64)
        f_true = np.asarray(data["f_true"], dtype=np.float64)
        f_residual = np.asarray(data["f_residual"], dtype=np.float64)

        row1 = {
            "Model": model_name,
            "L2 Error on Label Data": format_display(_extract_scalar(data, "u_fit_rel_l2")),
        }
        u_percentages = pointwise_relative_percentages(u_true, u_pred, u_thresholds)
        for threshold in u_thresholds:
            row1[threshold_header(threshold)] = format_percent(u_percentages[threshold])
        table1_rows.append(row1)

        row2 = {"Model": model_name}
        f_abs_percentages = residual_within_threshold_percentages(f_residual, f_thresholds)
        for threshold in f_thresholds:
            row2[tau_header(threshold)] = format_percent(f_abs_percentages[threshold])
        table2_rows.append(row2)

        row3 = {"Model": model_name}
        f_rel_percentages = normalized_residual_within_threshold_percentages(
            f_residual,
            f_true,
            f_relative_thresholds,
        )
        for threshold in f_relative_thresholds:
            row3[threshold_header(threshold)] = format_percent(f_rel_percentages[threshold])
        table3_rows.append(row3)

    csv_path = Path(args.csv).expanduser().resolve()
    excel_path = Path(args.excel).expanduser().resolve()
    if excel_path.suffix.lower() != ".xlsx":
        excel_path = excel_path.with_suffix(".xlsx")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=table1_fieldnames)
        writer.writeheader()
        writer.writerows(table1_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel_sheets(
        excel_path,
        [
            ("Table1", table1_rows, table1_fieldnames),
            ("Table2", table2_rows, table2_fieldnames),
            ("Table3", table3_rows, table3_fieldnames),
        ],
    )

    print_table(
        "Table 1: label-data L2 error and full-field u pointwise relative-error percentages",
        table1_rows,
        table1_fieldnames,
    )
    print_table(
        "Table 2: f residual percentages with |f_residual| <= tau",
        table2_rows,
        table2_fieldnames,
    )
    print_table(
        "Table 3: normalized f residual percentages with |f_residual| / |f_true| <= tau",
        table3_rows,
        table3_fieldnames,
    )
    print(f"Saved: {csv_path}")
    print(f"Saved: {excel_path}")


if __name__ == "__main__":
    main()
